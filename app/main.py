"""本地 AI 工作台 —— Web 应用入口。"""
import os
import time
import uuid
from functools import wraps

from flask import Flask, jsonify, request, send_file

from .config import load_settings
from .embeddings import get_embedder
from .llm import get_llm
from .logger import LogStore
from .rag import Index
from . import tools as tool_mod
from .auth import Auth
from .agent import Agent
from .orchestrator import Orchestrator
from .platforms import get_adapter

settings = load_settings()
# 统一转绝对路径：避免 Flask send_file 相对路径解析到 app/ 目录的坑
settings.data_dir = os.path.abspath(settings.data_dir)
settings.upload_dir = os.path.abspath(settings.upload_dir)
os.makedirs(settings.data_dir, exist_ok=True)
os.makedirs(settings.upload_dir, exist_ok=True)

embedder = get_embedder(settings)
index = Index(
    os.path.join(settings.data_dir, "index.json"),
    embedder=embedder,
    threshold=settings.retrieve_threshold,
    lexical_threshold=settings.retrieve_lexical_threshold,
    dedup_threshold=settings.dedup_threshold,
)
logs = LogStore(
    os.path.join(settings.data_dir, "logs.jsonl"),
    max_mb=settings.log_max_mb,
)
llm = get_llm(settings)
auth = Auth(settings, logs)
agent = Agent(settings, llm, logs, index)
orchestrator = Orchestrator(settings, llm, logs)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 单文件最大 20MB
_STATIC = os.path.join(os.path.dirname(__file__), "static")

SYSTEM_PROMPT = (
    "你是一个部署在客户本地、保护隐私的 AI 助手，请始终用中文回答。"
    "当用户提供参考资料时，优先依据参考资料回答；"
    "如果参考资料不足以回答，请明确说明。回答末尾列出引用的文档名。"
)


def _cost(tokens_in: int, tokens_out: int) -> float:
    return tokens_in / 1e6 * settings.price_in + tokens_out / 1e6 * settings.price_out


def _safe_output_file(name: str):
    """防路径穿越：只允许 data/output 下的文件，解析后校验真实路径。"""
    base = os.path.basename(str(name or ""))
    out_dir = os.path.realpath(os.path.join(settings.data_dir, "output"))
    real = os.path.realpath(os.path.join(out_dir, base))
    if real != out_dir and not real.startswith(out_dir + os.sep):
        return None
    if not os.path.exists(real):
        return None
    return real


def _as_bool(v, default: bool = False) -> bool:
    """严格布尔解析：仅接受真布尔或明确的 true/false 字符串，避免 "false" 被当真。"""
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes", "on")
    return default


def require_role(*roles):
    """后端角色校验：未启用登录验证=本地可信模式全部放行；启用后按角色拦截。"""
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not auth.auth_enabled():
                return fn(*args, **kwargs)
            token = _get_session_token()
            if not auth.check_session(token):
                return jsonify(error="未登录或会话已过期"), 401
            if auth.get_role(token) not in roles:
                return jsonify(error="无权限执行此操作"), 403
            return fn(*args, **kwargs)
        return wrapper
    return deco


# ---------------- 访问认证 ----------------

_AUTH_EXEMPT_PREFIXES = ("/api/auth",)


def _get_session_token():
    return request.cookies.get("session") or request.headers.get("X-Session", "")


@app.before_request
def require_auth():
    """除登录接口与首页外，全部接口要求有效会话；总开关关闭时全部拒绝。"""
    if request.path == "/" or request.path.startswith(_AUTH_EXEMPT_PREFIXES):
        return None
    if request.path.startswith("/api/health"):
        return None
    if not auth.access_enabled():
        return jsonify(error="演示已关闭，请联系管理员"), 403
    if auth.auth_enabled() and not auth.check_session(_get_session_token()):
        return jsonify(error="未登录或会话已过期"), 401


@app.post("/api/auth/request")
def auth_request():
    """发送验证码：需管理员口令（防他人刷邮件）。ADMIN_KEY 未设置时保持开放。"""
    key = request.headers.get("X-Admin-Key") or ""
    if settings.admin_key and key != settings.admin_key:
        return jsonify(error="口令错误"), 403
    result = auth.request_otp()
    return jsonify(result)


@app.post("/api/auth/login")
def auth_login():
    body = request.get_json(silent=True) or {}
    code = body.get("code") or ""
    ok, err = auth.verify_otp(code)
    if not ok:
        return jsonify(error=err), 401
    token = auth.create_session("client")
    resp = jsonify(ok=True, role="client")
    resp.set_cookie(
        "session", token,
        httponly=True, max_age=settings.session_ttl_seconds, samesite="Lax",
    )
    return resp


@app.post("/api/auth/login_key")
def auth_login_key():
    """口令登录。role: admin(ADMIN_KEY) / tech(TECH_KEY) / client(CLIENT_KEY)。"""
    if not auth.login_allowed():
        return jsonify(error="尝试过于频繁，请稍后再试"), 429
    body = request.get_json(silent=True) or {}
    key = (body.get("key") or "").strip()
    role = (body.get("role") or "").strip()
    if role == "admin":
        expected = settings.admin_key
    elif role == "tech":
        expected = settings.tech_key
    elif role == "client":
        expected = settings.client_key
    else:
        auth.login_fail()
        return jsonify(error="无效角色"), 400
    if not expected or key != expected:
        auth.login_fail()
        return jsonify(error="口令错误"), 403
    token = auth.create_session(role)
    resp = jsonify(ok=True, role=role)
    resp.set_cookie(
        "session", token,
        httponly=True, max_age=settings.session_ttl_seconds, samesite="Lax",
    )
    return resp


@app.get("/api/auth/check")
def auth_check():
    trial = auth.trial_seconds()
    token = _get_session_token()
    authed = auth.check_session(token)
    return jsonify(
        ok=authed,
        role=auth.get_role(token) if authed else "guest",
        auth_enabled=auth.auth_enabled(),
        access_enabled=auth.access_enabled(),
        trial_minutes=int(trial / 60) if trial else 0,
    )


@app.post("/api/auth/logout")
def auth_logout():
    auth.destroy_session(_get_session_token())
    resp = jsonify(ok=True)
    resp.delete_cookie("session")
    return resp


@app.get("/")
def home():
    return send_file(os.path.join(_STATIC, "index.html"))


@app.get("/api/health")
def health():
    return jsonify(status="ok", provider=llm.name, model=getattr(llm, "model", "-"))


@app.get("/api/config")
def config():
    return jsonify(
        provider=llm.name,
        model=getattr(llm, "model", "-"),
        retrieval=embedder.name,
        image_provider=settings.image_provider,
        docs=len(index.list()),
    )


@app.post("/api/chat")
def chat():
    body = request.get_json(silent=True) or {}
    msg = (body.get("message") or "").strip()
    if not msg:
        logs.add(type="chat_request", level="error", message="", error="消息不能为空")
        return jsonify(error="消息不能为空"), 400
    use_kb = body.get("use_kb", True)

    start = time.time()
    ret = index.retrieve(msg, settings.top_k) if use_kb else {"hits": [], "removed": 0}
    context = ret["hits"]
    logs.add(
        type="retrieval",
        message=msg[:60],
        hits=len(context),
        removed=ret["removed"],
        mode=embedder.name,
        top_score=round(context[0]["score"], 4) if context else 0,
    )

    user_prompt = msg
    if context:
        blocks = "\n\n".join(
            f"【来自文档《{c['doc']}》{('· ' + c['heading']) if c.get('heading') else ''}】\n{c['text']}"
            for c in context
        )
        user_prompt = f"参考资料：\n{blocks}\n\n问题：{msg}"

    try:
        answer, tokens_in, tokens_out = llm.chat(
            [{"role": "user", "content": user_prompt}], system=SYSTEM_PROMPT
        )
    except Exception as e:
        logs.add(type="chat_request", level="error", message=msg[:60], error=str(e)[:300])
        return jsonify(error=f"模型调用失败: {e}"), 502

    duration = (time.time() - start) * 1000
    logs.add(
        type="chat_request",
        message=msg[:60],
        hits=len(context),
        tokens_in=tokens_in,
        tokens_out=tokens_out,
        cost=round(_cost(tokens_in, tokens_out), 6),
        duration_ms=round(duration, 1),
        model=getattr(llm, "model", llm.name),
    )
    return jsonify(answer=answer, sources=context)


@app.post("/api/documents")
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="未选择文件"), 400
    name = os.path.basename(f.filename)
    raw = f.read()
    ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""
    if ext in ("docx", "pdf"):
        text = _extract_doc_text(ext, raw)
        if not text:
            return jsonify(error="无法从文件中提取文字（可能是扫描件/图片型 PDF）"), 400
    elif ext == "xlsx":
        try:
            text = _extract_xlsx_text(raw)
        except Exception as e:
            return jsonify(error=f"Excel 解析失败: {e}"), 400
        if not text.strip():
            return jsonify(error="Excel 中没有可读内容"), 400
    else:
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = raw.decode("gbk")
            except UnicodeDecodeError:
                return jsonify(error="仅支持文本(.txt/.md/.csv/.json/.log)、Word(.docx)、PDF(.pdf)、Excel(.xlsx)"), 400
    text = text.lstrip("\ufeff")  # 去掉 Windows 常见 BOM 头

    dest = os.path.join(settings.upload_dir, uuid.uuid4().hex + os.path.splitext(name)[1])
    with open(dest, "wb") as fp:
        fp.write(raw)

    doc = index.add_document(name, dest, text, settings.chunk_size, settings.chunk_min_size)
    logs.add(type="upload", name=name, chunks=len(doc["chunks"]), mode=embedder.name)
    return jsonify(id=doc["id"], name=doc["name"], chunks=len(doc["chunks"]))


def _extract_doc_text(ext: str, raw: bytes) -> str:
    """从 Word/PDF 提取文字（docx 解析段落，pdf 逐页提取）。"""
    try:
        import io
        if ext == "docx":
            from docx import Document
            doc = Document(io.BytesIO(raw))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return ""


def _extract_xlsx_text(raw: bytes, max_rows_per_sheet: int = 1000) -> str:
    """把 Excel 的【全部工作表】转成文本：工作表名作为标题，单元格用 | 分隔。

    例：
    # 工作表：中文目录
    商品名称 | 价格 | 描述
    夏季连衣裙 | 199 | 轻薄透气
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"# 工作表：{ws.title}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                parts.append("...(该表行数过多，已截断)")
                break
            cells = ["" if c is None else str(c).strip() for c in row]
            parts.append(" | ".join(cells).rstrip(" |") + "；")  # 行尾加分号，避免切片时行粘连
    wb.close()
    return "\n".join(p for p in parts if p.strip())


@app.get("/api/documents")
def documents():
    return jsonify(documents=index.list())


@app.get("/api/documents/<doc_id>")
def document_detail(doc_id: str):
    """查看单篇文档的切片详情（自检用）。"""
    info = index.detail(doc_id)
    if info is None:
        return jsonify(error="文档不存在"), 404
    return jsonify(info)


@app.post("/api/documents/reindex")
@require_role("admin", "tech")
def reindex_documents():
    """按当前 .env 切片参数重新切片全部文档（从原件重读）。"""
    n = index.reindex(settings.chunk_size, settings.chunk_min_size)
    logs.add(type="system", message=f"重新切片完成: {n} 篇")
    return jsonify(ok=True, docs=n)


@app.delete("/api/documents/<doc_id>")
@require_role("admin")
def delete_document(doc_id: str):
    doc = next((d for d in index.docs if d["id"] == doc_id), None)
    if not index.remove(doc_id):
        return jsonify(error="文档不存在"), 404
    logs.add(type="delete", name=(doc or {}).get("name", doc_id))
    return jsonify(ok=True)


# ---------------- 商品上架（平台对接） ----------------

@app.get("/api/listing/config")
def listing_config():
    adapter = get_adapter(settings.platform, settings)
    if adapter is None:
        return jsonify(platform=settings.platform, configured=False, message=f"未知平台 {settings.platform}")
    ok, msg = adapter.validate_config()
    return jsonify(platform=adapter.name, configured=ok, message=msg)


@app.post("/api/listing/upload")
def listing_upload():
    """批量上架：上传成品表格（或传 file_url），选列映射，发布到平台。"""
    adapter = get_adapter(settings.platform, settings)
    if adapter is None:
        return jsonify(error=f"未知平台 {settings.platform}"), 400

    raw = None
    name = None
    f = request.files.get("file")
    if f and f.filename:
        raw = f.read()
        name = os.path.basename(f.filename)
    else:
        file_url = (request.form.get("file_url") or "").strip()
        if file_url.startswith("/api/tools/download/"):
            fname = os.path.basename(file_url)
            p = _safe_output_file(fname)
            if p is None:
                return jsonify(error="文件不存在"), 404
            with open(p, "rb") as fh:
                raw = fh.read()
            name = fname
    if not raw:
        return jsonify(error="未提供成品文件"), 400

    try:
        rows = tool_mod._parse_table(name, raw)
    except Exception as e:
        return jsonify(error=f"表格解析失败: {e}"), 400
    if not rows:
        return jsonify(error="表格为空"), 400
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    ok, msg = adapter.validate_config()
    if not ok:
        return jsonify(error=msg), 400

    mapping = {
        "title_col": (request.form.get("title_col") or "").strip(),
        "body_col": (request.form.get("body_col") or "").strip(),
        "image_col": (request.form.get("image_col") or "").strip(),
        "tags_col": (request.form.get("tags_col") or "").strip(),
        "status": (request.form.get("status") or "draft").strip(),
    }
    result = adapter.upload_products(rows[1:], header, mapping, settings)
    logs.add(type="listing", platform=adapter.name, title_col=mapping["title_col"],
             created=result.get("created", 0), total=result.get("total", 0), level="info")
    return jsonify(result)


# ---------------- 编排器（多Agent分工流水线） ----------------

@app.post("/api/orchestrator/plan")
def orchestrator_plan():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="请上传表格文件(xlsx/csv)"), 400
    task = (request.form.get("task") or "").strip()
    if not task:
        return jsonify(error="请描述任务"), 400
    name = os.path.basename(f.filename)
    if not name.lower().endswith((".xlsx", ".csv")):
        return jsonify(error="仅支持 .xlsx / .csv"), 400
    raw = f.read()
    try:
        info = tool_mod.preview_table(name, raw)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    dest = os.path.join(settings.upload_dir, uuid.uuid4().hex + os.path.splitext(name)[1])
    with open(dest, "wb") as fp:
        fp.write(raw)
    try:
        plan_id, order = orchestrator.plan(task, name, dest, info["columns"])
    except Exception as e:
        return jsonify(error=f"编排失败: {e}"), 500
    # 预估生图费用提示
    est = None
    if order.get("images", {}).get("enabled"):
        n = info["total_rows"]
        est = {"count": n, "approx": f"约 ¥{round(n * 0.4, 1)}"}
    return jsonify(
        plan_id=plan_id,
        order=order,
        columns=info["columns"],
        total_rows=info["total_rows"],
        image_estimate=est,
    )


@app.post("/api/orchestrator/run")
def orchestrator_run():
    body = request.get_json(silent=True) or {}
    plan_id = (body.get("plan_id") or "").strip()
    approve_images = _as_bool(body.get("approve_images"))
    if not plan_id:
        return jsonify(error="缺少 plan_id"), 400
    try:
        result = orchestrator.run(plan_id, approve_images)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=f"执行失败: {e}"), 500
    return jsonify(result)


# ---------------- Agent 智能体 ----------------

@app.post("/api/agent/start")
def agent_start():
    body = request.get_json(silent=True) or {}
    task = (body.get("task") or "").strip()
    if not task:
        return jsonify(error="任务不能为空"), 400
    return jsonify(agent.start(task))


@app.post("/api/agent/confirm")
def agent_confirm():
    body = request.get_json(silent=True) or {}
    run_id = (body.get("run_id") or "").strip()
    approve = _as_bool(body.get("approve"))
    if not run_id:
        return jsonify(error="缺少 run_id"), 400
    return jsonify(agent.confirm(run_id, approve))


@app.post("/api/agent/cancel")
def agent_cancel():
    body = request.get_json(silent=True) or {}
    run_id = (body.get("run_id") or "").strip()
    if run_id:
        agent.cancel(run_id)
    return jsonify(ok=True)


# ---------------- 运行监控 ----------------

@app.get("/api/logs")
@require_role("admin", "tech")
def api_logs():
    limit = min(int(request.args.get("limit", 100)), 500)
    ftype = request.args.get("type") or None
    return jsonify(logs=logs.recent(limit, ftype))


@app.get("/api/stats")
@require_role("admin")
def api_stats():
    return jsonify(logs.stats(settings.price_in, settings.price_out))


@app.post("/api/logs/clear")
@require_role("admin")
def api_logs_clear():
    logs.clear()
    logs.add(type="system", message="日志已清空")
    return jsonify(ok=True)


# ---------------- 工具中心 ----------------

@app.post("/api/tools/text/batch")
def tool_text_batch():
    body = request.get_json(silent=True) or {}
    items = body.get("items") or []
    if not items:
        return jsonify(error="没有输入内容"), 400
    if len(items) > 200:
        return jsonify(error="单次最多 200 条"), 400
    results = tool_mod.batch_text(
        items,
        body.get("mode", "rewrite"),
        body.get("language", "英文"),
        body.get("instruction", ""),
        llm, logs, settings,
        concurrency=settings.tool_concurrency,
    )
    return jsonify(results=results)


@app.post("/api/tools/image/generate")
def tool_image_generate():
    body = request.get_json(silent=True) or {}
    prompts = body.get("prompts") or []
    if not prompts:
        return jsonify(error="没有输入提示词"), 400
    if len(prompts) > 20:
        return jsonify(error="单次最多 20 张"), 400
    images = tool_mod.batch_images(prompts, body.get("size", "512x512"), settings, logs)
    return jsonify(images=images, provider=settings.image_provider)


@app.get("/api/tools/image/<name>")
def tool_image_get(name: str):
    path = _safe_output_file(name)
    if path is None:
        return jsonify(error="文件不存在"), 404
    return send_file(path, mimetype="image/png")


@app.post("/api/tools/table/preview")
def tool_table_preview():
    f = request.files.get("file")
    if not f:
        return jsonify(error="未选择文件"), 400
    try:
        info = tool_mod.preview_table(os.path.basename(f.filename or ""), f.read())
    except ValueError as e:
        return jsonify(error=str(e)), 400
    return jsonify(info)


@app.post("/api/tools/table/process")
def tool_table_process():
    f = request.files.get("file")
    if not f:
        return jsonify(error="未选择文件"), 400
    name = os.path.basename(f.filename or "")
    raw = f.read()
    cols_raw = (request.form.get("columns") or request.form.get("column") or "").strip()
    columns = [c.strip() for c in cols_raw.split(",") if c.strip()]
    if not columns:
        return jsonify(error="未选择要处理的列"), 400
    mode = request.form.get("mode", "rewrite") or "rewrite"
    language = request.form.get("language", "英文") or "英文"
    instruction = request.form.get("instruction", "") or ""
    try:
        out_path, out_rows = tool_mod.process_table(
            name, raw, columns, mode, language, instruction, llm, logs, settings
        )
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=f"处理失败: {e}"), 500
    return jsonify(
        download="/api/tools/download/" + os.path.basename(out_path),
        rows=len(out_rows) - 1,
        preview=out_rows[:5],
    )


@app.get("/api/tools/download/<name>")
def tool_download(name: str):
    path = _safe_output_file(name)
    if path is None:
        return jsonify(error="文件不存在"), 404
    return send_file(path, as_attachment=True)


if __name__ == "__main__":
    app.run(host=settings.host, port=settings.port, threaded=True)
