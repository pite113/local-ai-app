"""工具中心：批量文本处理 / 批量作图 / 表格处理。

每个工具独立实现，互不影响；调用全部接入运行监控（token/成本/错误）。
"""
import base64
import csv
import io
import os
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed

import httpx

from .config import Settings
from .logger import LogStore


# ================= ① 批量文本处理 =================

def _build_prompt(mode: str, item: str, language: str, instruction: str) -> str:
    if mode == "translate":
        return f"请把下面的内容翻译成{language or '英文'}。只输出翻译结果：\n{item}"
    if mode == "rewrite":
        ins = f"。要求：{instruction}" if instruction else ""
        return f"请改写下面的内容，使其更专业、更有吸引力{ins}。只输出改写结果：\n{item}"
    ins = f"。要求：{instruction}" if instruction else ""
    return f"请根据下面的要点生成一段完整、有吸引力的文案{ins}。只输出文案本身：\n{item}"


def batch_text(items, mode, language, instruction, llm, logs: LogStore, settings: Settings,
               concurrency: int = 3):
    """批量处理：并发执行 + 相同内容自动去重（省 token），结果保持原顺序。"""
    system = (
        "你是一个专业的文案助手。严格按用户要求处理每一条文本，"
        "只输出处理结果本身，不要添加解释、编号、引号或多余换行。"
    )

    # 去重：相同 (mode, 内容) 只调用一次模型，结果映射回所有位置
    unique_items, key_to_uidx = [], {}
    for raw in items:
        item = (raw or "").strip()
        if not item:
            continue
        key = (mode, item)
        if key not in key_to_uidx:
            key_to_uidx[key] = len(unique_items)
            unique_items.append(item)

    def _process_one(uidx: int, item: str) -> str:
        prompt = _build_prompt(mode, item, language, instruction)
        start = time.time()
        try:
            answer, ti, to = llm.chat([{"role": "user", "content": prompt}], system=system)
            dur = (time.time() - start) * 1000
            logs.add(
                type="tool_text", mode=mode, item_index=uidx, message=item[:40],
                tokens_in=ti, tokens_out=to,
                cost=round(ti / 1e6 * settings.price_in + to / 1e6 * settings.price_out, 6),
                duration_ms=round(dur, 1), level="info",
            )
            return answer.strip()
        except Exception as e:
            logs.add(type="tool_text", mode=mode, item_index=uidx, message=item[:40],
                     level="error", error=str(e)[:200])
            return f"[处理失败] {e}"

    unique_results = [""] * len(unique_items)
    workers = max(1, min(int(concurrency), len(unique_items) or 1))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_process_one, i, it): i for i, it in enumerate(unique_items)}
        for fut in as_completed(futures):
            unique_results[futures[fut]] = fut.result()

    # 映射回原顺序（含空行占位）
    results = []
    for raw in items:
        item = (raw or "").strip()
        if not item:
            results.append("")
        else:
            results.append(unique_results[key_to_uidx[(mode, item)]])
    return results


# ================= ② 批量作图 =================

def _mock_image(prompt: str, size: str, out_path: str):
    """无 API 时的占位图（用于测试流程），API 配置好后自动替换。"""
    from PIL import Image, ImageDraw, ImageFont
    w = {"512x512": 512, "1024x1024": 1024}.get(size, 512)
    img = Image.new("RGB", (w, w), (30, 144, 255))
    draw = ImageDraw.Draw(img)
    # 用系统中文字体渲染，避免中文变方框
    font = None
    for fp in (r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\simsun.ttc"):
        if os.path.exists(fp):
            try:
                font = ImageFont.truetype(fp, 22)
                break
            except Exception:
                continue
    lines, cur = [], ""
    for ch in prompt:
        cur += ch
        if len(cur) >= 16:
            lines.append(cur)
            cur = ""
    if cur:
        lines.append(cur)
    y = w // 2 - min(len(lines), 10) * 26
    for ln in lines[:10]:
        draw.text((20, y), ln, fill=(255, 255, 255), font=font)
        y += 26
    img.save(out_path, "PNG")


def _api_image(prompt: str, size: str, settings: Settings, out_path: str):
    """OpenAI 兼容 /images/generations 接口（硅基流动等）。"""
    headers = {"Authorization": f"Bearer {settings.image_api_key}"}
    payload = {"model": settings.image_api_model, "prompt": prompt, "size": size, "n": 1}
    r = httpx.post(
        settings.image_api_base.rstrip("/") + "/images/generations",
        json=payload, headers=headers, timeout=180,
    )
    r.raise_for_status()
    data = r.json()["data"][0]
    if data.get("b64_json"):
        with open(out_path, "wb") as f:
            f.write(base64.b64decode(data["b64_json"]))
    elif data.get("url"):
        img = httpx.get(data["url"], timeout=120)
        img.raise_for_status()
        with open(out_path, "wb") as f:
            f.write(img.content)
    else:
        raise RuntimeError("图片接口返回格式无法识别")


def batch_images(prompts, size: str, settings: Settings, logs: LogStore):
    out_dir = os.path.join(settings.data_dir, "output")
    os.makedirs(out_dir, exist_ok=True)
    results = []
    for idx, p in enumerate(prompts):
        p = (p or "").strip()
        if not p:
            continue
        fname = f"img_{uuid.uuid4().hex}.png"
        fpath = os.path.join(out_dir, fname)
        try:
            if settings.image_provider == "api" and settings.image_api_key:
                _api_image(p, size, settings, fpath)
            else:
                _mock_image(p, size, fpath)
            logs.add(type="tool_image", message=p[:40], file=fname,
                     provider=settings.image_provider, level="info")
            results.append({"file": fname, "url": f"/api/tools/image/{fname}", "prompt": p[:60]})
        except Exception as e:
            logs.add(type="tool_image", message=p[:40], level="error", error=str(e)[:200])
            results.append({"file": None, "url": None, "prompt": p[:60], "error": str(e)})
    return results


# ================= ③ 表格处理 =================

def _parse_table(filename: str, raw: bytes):
    if filename.lower().endswith(".csv"):
        text = None
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            raise ValueError("CSV 编码无法识别，请用 UTF-8 或 GBK 保存")
        return list(csv.reader(io.StringIO(text)))
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def preview_table(filename: str, raw: bytes):
    rows = _parse_table(filename, raw)
    if not rows:
        raise ValueError("表格为空")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return {"columns": header, "preview": rows[:6], "total_rows": max(len(rows) - 1, 0)}


def process_table(filename, raw, columns, mode, language, instruction, llm, logs, settings):
    """多列批量处理：每列生成一个"处理结果_列名"新列。columns 为列名列表。"""
    rows = _parse_table(filename, raw)
    if not rows:
        raise ValueError("表格为空")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    missing = [c for c in columns if c not in header]
    if missing:
        raise ValueError(f"找不到列「{'、'.join(missing)}」，现有列：{'、'.join(header)}")
    col_idxs = [header.index(c) for c in columns]

    # 逐列收集待处理文本（跨列不去重，列内 batch_text 自动去重）
    col_results = []
    for ci in col_idxs:
        items = [
            str(r[ci]).strip() if ci < len(r) and r[ci] is not None else ""
            for r in rows[1:]
        ]
        col_results.append(batch_text(items, mode, language, instruction, llm, logs, settings))

    new_header = header + [f"处理结果_{c}" for c in columns]
    out_rows = [new_header]
    for i, row in enumerate(rows[1:]):
        out_rows.append(list(row) + [col_results[j][i] for j in range(len(columns))])

    out_name = f"processed_{uuid.uuid4().hex}{os.path.splitext(filename)[1]}"
    out_path = os.path.join(settings.data_dir, "output", out_name)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    if filename.lower().endswith(".csv"):
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(out_rows)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in out_rows:
            ws.append(row)
        wb.save(out_path)
    logs.add(type="tool_table", name=filename, columns=",".join(columns), mode=mode,
             rows=len(out_rows) - 1, level="info")
    return out_path, out_rows
